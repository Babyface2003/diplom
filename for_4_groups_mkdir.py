import os

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Font

def split_data_by_condition(data):
    files = []
    current_file_data = []

    for idx, row in data.iterrows():
        if pd.isna(row.iloc[0]) and pd.notna(row.iloc[1]):
            if current_file_data:
                files.append(pd.DataFrame(current_file_data))
                current_file_data = []
        current_file_data.append(row)

    if current_file_data:  # Добавляем оставшиеся данные
        files.append(pd.DataFrame(current_file_data))

    return files

# Основная функция обработки
def process_files_in_directory(directory):
    for root, dirs, files in os.walk(directory):
        for file_name in files:
            file_path = os.path.join(root, file_name)

            # Проверяем, является ли файл Excel
            if file_name.endswith(".xlsx"):
                workbook = load_workbook(file_path)
                sheet_name = workbook.sheetnames[0]  # Предполагаем, что лист один
                sheet = workbook[sheet_name]

                # Считываем данные вместе со стилями
                data = []
                styles = []

                for row in sheet.iter_rows():
                    row_data = []
                    row_styles = []
                    for cell in row:
                        row_data.append(cell.value)
                        row_styles.append(cell.font.bold)  # Сохраняем информацию о жирном шрифте
                    data.append(row_data)
                    styles.append(row_styles)

                # Преобразуем данные в DataFrame
                data_df = pd.DataFrame(data).dropna(how='all')

                # Разделяем данные по критерию
                split_files = split_data_by_condition(data_df)

                if len(split_files) > 1:
                    base_name = os.path.splitext(file_name)[0]

                    for i, df in enumerate(split_files):
                        output_path = os.path.join(root, f"{base_name}_{i + 1}.xlsx")

                        # Создаем новый файл Excel с сохранением стилей
                        new_workbook = Workbook()
                        new_sheet = new_workbook.active

                        for row_idx, row_data in enumerate(df.values, start=1):
                            for col_idx, value in enumerate(row_data, start=1):
                                cell = new_sheet.cell(row=row_idx, column=col_idx, value=value)

                                # Применяем сохраненный стиль (жирный шрифт)
                                if len(styles) > row_idx - 1 and len(styles[row_idx - 1]) > col_idx - 1:
                                    if styles[row_idx - 1][col_idx - 1]:  # Если был жирным
                                        cell.font = Font(bold=True)

                        new_workbook.save(output_path)

                    # Удаляем исходный файл
                    if os.path.exists(file_path):
                        os.remove(file_path)
                        print(f"Исходный файл '{file_path}' был удалён.")

                else:
                    print(f"Файл '{file_name}' не требует разделения.")

# Список директорий для обработки
base_dir = r"C:\Users\ivanr\PycharmProjects\VKR"
sub_dirs = ["1_курс", "2_курс", "3_курс", "4_курс", "5_курс", "6_курс"]

# Обрабатываем все указанные директории
for sub_dir in sub_dirs:
    dir_path = os.path.join(base_dir, sub_dir)
    if os.path.exists(dir_path):
        process_files_in_directory(dir_path)

print("Обработка завершена.")