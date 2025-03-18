import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
from openpyxl.cell import Cell


# Путь к основной директории
base_path = r"C:\Users\ivanr\PycharmProjects\VKR"
subfolders = ["1_курс", "2_курс", "3_курс", "4_курс", "5_курс", "6_курс"]

# Ключевые слова для поиска названий групп
group_keywords = ['МДС', 'ИДБ', 'ЭДБ', 'АДБ', 'МДБ']

# Проход по всем поддиректориям
for subfolder in subfolders:
    folder_path = os.path.join(base_path, subfolder)

    # Проверка наличия папки
    if not os.path.exists(folder_path):
        print(f"Папка {folder_path} не найдена, пропускаем.")
        continue

    # Поиск всех файлов Excel в папке
    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.xlsx'):
                file_path = os.path.join(root, file)
                print(f"Обработка файла: {file_path}")

                try:
                    # Загрузка файла Excel
                    wb = load_workbook(file_path)
                    ws = wb.active

                    # Поиск уникальных названий групп
                    header = [cell.value if isinstance(cell, Cell) else cell for cell in ws[1]]  # Заголовок первой строки
                    group_columns = [col for col in header if
                                     isinstance(col, str) and any(key in col for key in group_keywords)]

                    # Папка для сохранения результатов
                    output_folder = os.path.dirname(file_path)

                    # Обработка каждой группы
                    for group in group_columns:
                        group_index = header.index(group)  # Индекс столбца группы
                        group_cols = [group_index, group_index + 1, group_index + 2]  # Столбцы ФИО группы

                        # Извлечение данных группы
                        group_data = []
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                            # Проверяем, есть ли данные в столбцах группы
                            if any(cell.value for cell in [row[i] for i in group_cols]):
                                group_data.append(row)

                        # Пропуск, если данных нет
                        if not group_data:
                            continue

                        # Создание новой книги для группы
                        new_wb = Workbook()
                        new_ws = new_wb.active
                        new_ws.title = group
                        new_ws.append(['№', 'Фамилия', 'Имя', 'Отчество'])

                        # Добавление данных с нумерацией и копированием стилей
                        for idx, row in enumerate(group_data, start=1):
                            # Извлекаем значения из столбцов группы
                            new_row = [idx] + [row[i].value for i in group_cols]
                            new_ws.append(new_row)

                            # Копирование стилей
                            for source_idx, target_cell in zip(group_cols, new_ws[idx + 1][1:]):  # Пропускаем номер
                                source_cell = row[source_idx]
                                target_cell.font = copy(source_cell.font)
                                target_cell.alignment = copy(source_cell.alignment)
                                target_cell.border = copy(source_cell.border)
                                target_cell.fill = copy(source_cell.fill)

                        # Сохранение новой книги
                        output_path = os.path.join(output_folder, f"{group}.xlsx")
                        new_wb.save(output_path)
                        print(f"Данные для группы {group} сохранены в файл {output_path}")

                    # Удаление исходного файла после обработки
                    os.remove(file_path)
                    print(f"Файл {file_path} удалён после обработки.")

                except Exception as e:
                    print(f"Ошибка при обработке файла {file_path}: {e}")
